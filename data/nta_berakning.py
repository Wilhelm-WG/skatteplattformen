"""
Skatteplattformen — Beräkningsmodell v1.0
Schablonberäkning av offentliga förmåner per åldersprofil
Källa: ESV 2024, SKR 2024, SCB KPI, Försäkringskassan, CSN
"""
import os
import sys
from openpyxl import load_workbook
_DIR = os.path.dirname(os.path.abspath(__file__))
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# ── Constants ─────────────────────────────────────────────────────────────────
BEFOLKNING   = 10_550_000
BASÅR        = 2024

def pc_tkr(mdkr): 
    """Mdkr → tkr per capita"""
    return round(mdkr * 1_000_000 / BEFOLKNING, 1)

# ── Per capita bastal 2024 (tkr/år, alla åldrar genomsnitt) ──────────────────
PC = {
    "sjukvård":    pc_tkr(102.4 + 490.0),  # 56.2 — UO9 + SKR regionsjukvård
    "sjukforsakr": pc_tkr(131.5),           # 12.5 — UO10
    "pension_g":   pc_tkr(72.1),            # 6.8  — UO11 garantipension
    "familj":      pc_tkr(107.5),           # 10.2 — UO12 barnbidrag/föräldraförs.
    "studiemedel": pc_tkr(29.4),            # 2.8  — UO15 CSN
    "hogskola":    pc_tkr(100.2),           # 9.5  — UO16
    "skola_k":     pc_tkr(370.0),           # 35.1 — Kommunal grundskola/gymn (SKR)
    "infra":       pc_tkr(82.1),            # 7.8  — UO22 kommunikationer
    "kom_bidrag":  pc_tkr(164.7),           # 15.6 — UO25 generella bidrag
    "aldreomsorg": pc_tkr(145.0),           # 13.7 — Kommunal äldreomsorg (SKR)
}

# ── Åldersviktningskoefficienter ─────────────────────────────────────────────
# Normaliserade: genomsnitt = 1.0 per post (bevarar total)
# Källor: Socialstyrelsen, Försäkringskassan, CSN, UKÄ, SKR, Trafikverket RVU

RAW_VIKTER = {
    #                         18å   25å   35å   45å   65å
    "sjukvård":    "källa: Socialstyrelsen öppna jämförelser, kostnad per åldersgrupp",
    "sjukforsakr": "källa: Försäkringskassan årsredovisning, sjukpenningdagar per åldersgrupp",
    "pension_g":   "källa: Pensionsmyndigheten, garantipension endast >65 år",
    "familj":      "källa: SCB barnfamiljeandel per åldersgrupp, barnbidrag 0–16 år",
    "studiemedel": "källa: CSN årsstatistik, studerande per åldersgrupp",
    "hogskola":    "källa: UKÄ årsrapport, registrerade studenter per åldersgrupp",
    "skola_k":     "källa: SCB elever per skolform, SCB befolkningsstatistik 6–19 år",
    "infra":       "källa: Trafikverket RVU Sverige, resa per åldersgrupp",
    "kom_bidrag":  "källa: SKR, kommunal service intensitet per åldersgrupp",
    "aldreomsorg": "källa: Socialstyrelsen, mottagare äldreomsorg per åldersgrupp",
}

# Oinormaliserade vikter — metodologiskt motiverade
VIKTER_RAW = {
    "sjukvård":    [0.28, 0.35, 0.62, 1.18, 3.40],  
    "sjukforsakr": [0.05, 0.30, 0.70, 1.40, 0.45],  
    "pension_g":   [0.00, 0.00, 0.00, 0.00, 3.20],  
    "familj":      [0.10, 0.90, 1.80, 0.90, 0.10],  
    "studiemedel": [2.50, 2.80, 0.20, 0.05, 0.00],  
    "hogskola":    [1.20, 1.60, 0.50, 0.30, 0.10],  
    "skola_k":     [3.50, 0.80, 1.20, 0.80, 0.05],  
    "infra":       [0.70, 0.90, 1.10, 1.10, 0.80],  
    "kom_bidrag":  [0.50, 0.70, 0.90, 0.95, 1.80],  
    "aldreomsorg": [0.00, 0.00, 0.00, 0.05, 4.50],  
}

# Normalisera: sum(vikter) = 5 (en per åldersgrupp, genomsnitt = 1)
VIKTER = {}
for key, vals in VIKTER_RAW.items():
    s = sum(vals)
    if s > 0:
        VIKTER[key] = [round(v / s * 5, 4) for v in vals]
    else:
        VIKTER[key] = [0.0] * 5

AGE_PROFILES = [
    {"label": "18 år", "age": 18, "idx": 0},
    {"label": "25 år", "age": 25, "idx": 1},
    {"label": "35 år", "age": 35, "idx": 2},
    {"label": "45 år", "age": 45, "idx": 3},
    {"label": "65 år", "age": 65, "idx": 4},
]

POSTER_DEF = [
    ("Sjukvård (stat + region)",     "sjukvård",    "UO9 + SKR regionsjukvård"),
    ("Sjukförsäkring",               "sjukforsakr", "UO10 ekonomisk trygghet sjukdom"),
    ("Barnbidrag & föräldraförs.",   "familj",      "UO12 familjer och barn"),
    ("Studiemedel (CSN)",            "studiemedel", "UO15 studiestöd"),
    ("Högskoleutbildning",           "hogskola",    "UO16 utbildning & forskning"),
    ("Grundskola & gymnasium",       "skola_k",     "SKR kommunal skola"),
    ("Vägar & järnvägar",            "infra",       "UO22 kommunikationer"),
    ("Kommunal service",             "kom_bidrag",  "UO25 + SKR kommunal service"),
    ("Äldreomsorg",                  "aldreomsorg", "SKR kommunal äldreomsorg"),
    ("Garantipension",               "pension_g",   "UO11 ekonomisk trygghet ålderdom"),
]

# ── Beräkna årsförmåner per profil ───────────────────────────────────────────
def berakna_arsformaaner(age_idx):
    res = {}
    total = 0.0
    for namn, key, _ in POSTER_DEF:
        val = round(PC[key] * VIKTER[key][age_idx], 1)
        res[namn] = val
        total += val
    res["TOTALT"] = round(total, 1)
    return res

# ── Livstidsmodell — läser från livsfaser_v2.json (enda sanningskällan) ───────
# Kategorimap: JSON 5-kategorier → Python 10-poster (proportionell fördelning)
# Uppdatera ALLTID livsfaser_v2.json om värdena behöver ändras.

def _ladda_livsfaser():
    """Läser livsfaser_v2.json och konverterar till Python-modellens format."""
    livsfaser_path = os.path.join(_DIR, 'livsfaser_v2.json')
    with open(livsfaser_path, encoding='utf-8') as f:
        data = json.load(f)

    livsfaser = []
    for fas in data['faser']:
        fv  = fas['förmåner_tkr_per_år']
        utb = fv['utbildning']['värde']
        sj  = fv['sjukvård']['värde']
        soc = fv['socialt_skydd']['värde']
        koll= fv['kollektivt']['värde']
        oms = fv['äldreomsorg']['värde']

        # Fördelning utbildning → Python-poster baserat på fas
        ålder_mitt = (fas['min_ålder'] + fas['max_ålder']) / 2
        if ålder_mitt <= 5:
            sk = utb;          hg = 0;         sm = 0
        elif ålder_mitt <= 15:
            sk = utb * 0.95;  hg = 0;         sm = utb * 0.05
        elif ålder_mitt <= 18:
            sk = utb * 0.92;  hg = 0;         sm = utb * 0.08
        elif ålder_mitt <= 25:
            sk = 0;            hg = utb * 0.74; sm = utb * 0.26
        else:
            sk = utb * 0.30;  hg = utb * 0.50; sm = utb * 0.20

        # Fördelning socialt skydd → sjukforsakr + familj + pension_g
        if ålder_mitt >= 65:
            sf = soc * 0.10; fj = 0;          pg = soc * 0.90
        elif ålder_mitt >= 55:
            sf = soc * 0.60; fj = soc * 0.05; pg = 0
        elif ålder_mitt >= 35:
            sf = soc * 0.35; fj = soc * 0.65; pg = 0
        elif ålder_mitt >= 20:
            sf = soc * 0.25; fj = soc * 0.75; pg = 0
        else:
            sf = 0;           fj = soc;        pg = 0

        # Kollektivt → infra 50% + kom_bidrag 50%
        inf = koll * 0.50
        kb  = koll * 0.50

        livsfaser.append((
            fas['min_ålder'], fas['max_ålder'],
            {
                "sjukvård":    round(sj,  1),
                "sjukforsakr": round(sf,  1),
                "familj":      round(fj,  1),
                "studiemedel": round(sm,  1),
                "hogskola":    round(hg,  1),
                "skola_k":     round(sk,  1),
                "infra":       round(inf, 1),
                "kom_bidrag":  round(kb,  1),
                "aldreomsorg": round(oms, 1),
                "pension_g":   round(pg,  1),
            }
        ))
    return livsfaser

LIVSFASER = _ladda_livsfaser()

def livstid_formaaner(target_age):
    """Ackumulerade förmåner tkr (2024 priser) från födseln t.o.m. target_age"""
    total_by_post = {n: 0.0 for n,_,_ in POSTER_DEF}
    total_by_post["Barnomsorgsåren"] = 0.0
    grand_total = 0.0
    for (start, end, årsvals) in LIVSFASER:
        years = max(0, min(end, target_age) - start + 1)
        if years <= 0: continue
        for key, val in årsvals.items():
            for namn, k, _ in POSTER_DEF:
                if k == key:
                    total_by_post[namn] = total_by_post.get(namn, 0) + val * years
                    grand_total += val * years
                    break
    return round(grand_total), total_by_post

# ── Skatteberäkning ───────────────────────────────────────────────────────────
def berakna_skatt_livet(target_age, bruttolön_tkr=500):
    """Total skatt betald från 20 år (inkl. arbetsgivaravgifter), tkr 2024 priser"""
    arbetsår = max(0, min(target_age, 65) - 20)  # Arbetar 20–65
    kommunalskatt  = bruttolön_tkr * 0.3200  # 32% snitt
    statlig_skatt  = max(0, bruttolön_tkr - 598) * 0.20
    arbgivaravg    = bruttolön_tkr * 0.3142
    moms           = (bruttolön_tkr - kommunalskatt - statlig_skatt) * 0.065
    total_per_år   = kommunalskatt + statlig_skatt + arbgivaravg + moms
    return round(total_per_år * arbetsår)

# ── EXCEL: Lägg till ny flik ──────────────────────────────────────────────────
import os
_EXCEL = os.path.join(_DIR, '..', 'skatteplattformen_data.xlsx')
wb = load_workbook(_EXCEL)

if "Beräkningsmodell" in wb.sheetnames:
    del wb["Beräkningsmodell"]

ws = wb.create_sheet("Beräkningsmodell", 3)
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "1B4D3E"

def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def hf(sz=10, bold=True, col="FFFFFF"): return Font(name="Arial", size=sz, bold=bold, color=col)
def bf(sz=10, bold=False, col="1A1814"): return Font(name="Arial", size=sz, bold=bold, color=col)
def gcl(ci): return get_column_letter(ci)

# ─── ROW 1: Titel ─────────────────────────────────────────────────────────────
ws.merge_cells("A1:R1")
ws["A1"] = "Skatteplattformen — Beräkningsmodell v1.0 | Schablonförmåner per åldersprofil (tkr, 2024 priser)"
ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
ws["A1"].fill = fill("1B4D3E")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:R2")
ws["A2"] = "Källa: ESV 2025:8, SKR 2024, SCB KPI, Försäkringskassan, CSN, UKÄ | Metod: viktade genomsnitt per livsfas | Alla belopp i 2024 års priser"
ws["A2"].font = Font(name="Arial", size=9, color="555555")
ws["A2"].fill = fill("F4F1EB")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 18

# ─── SEKTION A: Per capita bastal ─────────────────────────────────────────────
r = 4
ws.merge_cells(f"A{r}:R{r}")
ws[f"A{r}"] = "A. PER CAPITA-BASTAL 2024 (alla åldrar, obeviktat genomsnitt)"
ws[f"A{r}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ws[f"A{r}"].fill = fill("2E6DA4")
ws[f"A{r}"].alignment = Alignment(indent=1, vertical="center")
ws.row_dimensions[r].height = 24
r += 1

for h, w in [("Post",35),("Belopp (Mdkr)",16),("Per capita (tkr/år)",20),("Källa / UO",45),("Metodnot",50)]:
    ci = ["Post","Belopp (Mdkr)","Per capita (tkr/år)","Källa / UO","Metodnot"].index(h)+1
    c = ws.cell(row=r, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill("1A3557")
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 26
r += 1

PC_SOURCES = [
    ("Sjukvård (stat + region)", 102.4+490.0, PC["sjukvård"],    "UO9 + SKR regionernas sjukvård",    "SKR: 88% av regionkostnad = sjukvård. Stat+region = fullständig bild."),
    ("Sjukförsäkring",           131.5,       PC["sjukforsakr"], "UO10 Ekon. trygghet vid sjukdom",   "Sjukpenning, sjukersättning, rehab. FK årsredovisning 2024."),
    ("Garantipension",           72.1,        PC["pension_g"],   "UO11 Ekon. trygghet vid ålderdom",  "Enbart garantipension. Inkomstpension = avgiftsfinansierad, ingår ej."),
    ("Barnbidrag & föräldraförs.",107.5,      PC["familj"],      "UO12 Ekon. trygghet familjer/barn", "Barnbidrag, föräldrapenning, underhållsstöd, aktivitetsstöd."),
    ("Studiemedel (CSN)",        29.4,        PC["studiemedel"], "UO15 Studiestöd",                   "CSN studiemedel, studiebidrag. CSN årsredovisning 2024."),
    ("Högskoleutbildning",       100.2,       PC["hogskola"],    "UO16 Utbildning & universitetsforskning", "Statsanslag till lärosäten. Ej kommunal vuxenutbildning."),
    ("Grundskola & gymnasium",   370.0,       PC["skola_k"],     "SKR kommunal skola",                "SKR sektorn i siffror 2024. Kommunernas kostnad grund-/gymnasieskola."),
    ("Vägar & järnvägar",        82.1,        PC["infra"],       "UO22 Kommunikationer",              "Trafikverket väg+banhållning. Ej privata vägar."),
    ("Kommunal service",         164.7,       PC["kom_bidrag"],  "UO25 Allmänna bidrag till kommuner","Generellt statsbidrag. Finansierar bl.a. sophantering, bibliotek, LSS."),
    ("Äldreomsorg",              145.0,       PC["aldreomsorg"], "SKR kommunal äldreomsorg",          "SKR 2024. Hemtjänst, särskilt boende. Ej privata utförare utan off. finansiering."),
]

total_mdkr = 0; total_pc = 0
for namn, mdkr, pc_val, källa, not_ in PC_SOURCES:
    bg = "F4F1EB" if (r % 2 == 0) else "FFFFFF"
    for ci, val in enumerate([namn, mdkr, pc_val, källa, not_], 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = bf(10)
        c.fill = fill(bg)
        c.alignment = Alignment(vertical="center", wrap_text=(ci in [1,4,5]))
        if ci == 2: c.number_format = '#,##0.0'
        if ci == 3: c.number_format = '#,##0.0'
    ws.row_dimensions[r].height = 28
    total_mdkr += mdkr; total_pc += pc_val
    r += 1

# Totalsummarad
for ci, val in enumerate(["TOTALT", total_mdkr, total_pc, "ESV+SKR+SCB", ""], 1):
    c = ws.cell(row=r, column=ci, value=val)
    c.font = bf(10, bold=True, col="FFFFFF")
    c.fill = fill("1B4D3E")
    if ci == 2: c.number_format = '#,##0.0'
    if ci == 3: c.number_format = '#,##0.0'
    c.alignment = Alignment(vertical="center")
ws.row_dimensions[r].height = 24; r += 2

# ─── SEKTION B: Viktningskoefficienter ────────────────────────────────────────
ws.merge_cells(f"A{r}:R{r}")
ws[f"A{r}"] = "B. ÅLDERSVIKTNINGSKOEFFICIENTER (normaliserade: genomsnitt = 1.0)"
ws[f"A{r}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ws[f"A{r}"].fill = fill("2E6DA4")
ws[f"A{r}"].alignment = Alignment(indent=1, vertical="center")
ws.row_dimensions[r].height = 24; r += 1

vikt_hdrs = ["Post", "18 år", "25 år", "35 år", "45 år", "65 år", "Viktkälla"]
vikt_widths = [30, 9, 9, 9, 9, 9, 55]
for ci, (h, w) in enumerate(zip(vikt_hdrs, vikt_widths), 1):
    c = ws.cell(row=r, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill("1A3557")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[gcl(ci)].width = w
ws.row_dimensions[r].height = 26; r += 1

VIKT_NOTER = {
    "sjukvård":    "Socialstyrelsen Öppna jämförelser, hälso- och sjukvårdskostnad per åldersgrupp",
    "sjukforsakr": "Försäkringskassan årsredovisning 2024, sjukpenningdagar per åldersgrupp",
    "pension_g":   "Enbart >65 år. Pensionsmyndigheten statistikrapport 2024.",
    "familj":      "SCB befolkningsstatistik barnfamiljer + FK föräldrapenningstatistik",
    "studiemedel": "CSN årsstatistik studerande per åldersklass",
    "hogskola":    "UKÄ årsrapport 2024, registrerade studenter per åldersgrupp",
    "skola_k":     "SCB elever per skolform; obligatorisk skola 6–15 år, gymnasium 16–19 år",
    "infra":       "Trafikverket RVU Sverige 2020–2021, resor per person och åldersgrupp",
    "kom_bidrag":  "SKR 2024, kommunal service fördelning; äldreomsorg tyngre >65",
    "aldreomsorg": "Socialstyrelsen 2024, andel mottagare av äldreomsorg per åldersgrupp",
}

for namn, key, _ in POSTER_DEF:
    bg = "F4F1EB" if (r % 2 == 0) else "FFFFFF"
    vikter = VIKTER.get(key, [0]*5)
    row_vals = [namn] + vikter + [VIKT_NOTER.get(key, "")]
    for ci, val in enumerate(row_vals, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = bf(10)
        c.fill = fill(bg)
        c.alignment = Alignment(vertical="center", wrap_text=(ci in [1,7]))
        if ci in [2,3,4,5,6]:
            c.number_format = '0.000'
            # Colour-code: high weight = darker green
            if isinstance(val, float) and val > 2.0:
                c.fill = fill("D4EBE0")
                c.font = bf(10, bold=True, col="0F6E56")
    ws.row_dimensions[r].height = 30; r += 1
r += 1

# ─── SEKTION C: Årsförmåner per profil ────────────────────────────────────────
ws.merge_cells(f"A{r}:R{r}")
ws[f"A{r}"] = "C. BERÄKNADE ÅRSFÖRMÅNER PER ÅLDERSPROFIL (tkr/år, 2024 priser) = A × B"
ws[f"A{r}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ws[f"A{r}"].fill = fill("2E6DA4")
ws[f"A{r}"].alignment = Alignment(indent=1, vertical="center")
ws.row_dimensions[r].height = 24; r += 1

age_hdrs = ["Post (tkr/år, 2024 priser)", "18 år", "25 år", "35 år", "45 år", "65 år", "Per capita snitt"]
for ci, h in enumerate(age_hdrs, 1):
    c = ws.cell(row=r, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill("1A3557")
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 26; r += 1

# Beräkna
age_totals = [0.0] * 5
for namn, key, _ in POSTER_DEF:
    bg = "F4F1EB" if (r % 2 == 0) else "FFFFFF"
    ws.cell(row=r, column=1, value=namn).font = bf(10)
    ws.cell(row=r, column=1).fill = fill(bg)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
    vals = []
    for i in range(5):
        val = round(PC[key] * VIKTER[key][i], 1)
        vals.append(val)
        age_totals[i] += val
        c = ws.cell(row=r, column=i+2, value=val)
        c.font = bf(10)
        c.fill = fill(bg)
        c.number_format = '#,##0.0'
        c.alignment = Alignment(horizontal="right", vertical="center")
    # Snitt
    c = ws.cell(row=r, column=7, value=PC[key])
    c.font = bf(10, col="2E6DA4")
    c.fill = fill(bg)
    c.number_format = '#,##0.0'
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[r].height = 24; r += 1

# Total rad
ws.cell(row=r, column=1, value="TOTALT ÅRSFÖRMÅNER").font = bf(10, bold=True, col="FFFFFF")
ws.cell(row=r, column=1).fill = fill("1B4D3E")
for i, tot in enumerate(age_totals):
    c = ws.cell(row=r, column=i+2, value=round(tot, 1))
    c.font = bf(10, bold=True, col="FFFFFF")
    c.fill = fill("1B4D3E")
    c.number_format = '#,##0.0'
    c.alignment = Alignment(horizontal="right", vertical="center")
c = ws.cell(row=r, column=7, value=round(sum(PC.values()), 1))
c.font = bf(10, bold=True, col="FFFFFF")
c.fill = fill("1B4D3E")
c.number_format = '#,##0.0'
c.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[r].height = 24; r += 2

# ─── SEKTION D: Livstidsförmåner ──────────────────────────────────────────────
ws.merge_cells(f"A{r}:R{r}")
ws[f"A{r}"] = "D. LIVSTIDSFÖRMÅNER FRÅN FÖDSELN (tkr, 2024 priser) — ackumulerat per åldersprofil"
ws[f"A{r}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ws[f"A{r}"].fill = fill("2E6DA4")
ws[f"A{r}"].alignment = Alignment(indent=1, vertical="center")
ws.row_dimensions[r].height = 24; r += 1

life_hdrs = ["Livsfas", "Årsintervall", "18 år", "25 år", "35 år", "45 år", "65 år", "Metodnot"]
for ci, h in enumerate(life_hdrs, 1):
    c = ws.cell(row=r, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill("1A3557")
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 26; r += 1

LIVSFAS_LABELS = [
    "Spädbarn & förskola",
    "Grundskola",
    "Gymnasium",
    "Högskolestudier",
    "Barnfamilj-fas",
    "Mitt-karriär",
    "Senior (45–64)",
    "Pensionär (65–80)",
]

col_ages = [18, 25, 35, 45, 65]
life_totals = [0.0] * 5

for fi, ((start, end, årsvals), lbl) in enumerate(zip(LIVSFASER, LIVSFAS_LABELS)):
    bg = "F4F1EB" if (r % 2 == 0) else "FFFFFF"
    ws.cell(row=r, column=1, value=lbl).font = bf(10); ws.cell(row=r,column=1).fill=fill(bg); ws.cell(row=r,column=1).alignment=Alignment(vertical="center")
    ws.cell(row=r, column=2, value=f"{start}–{end} år").font = bf(9, col="666666"); ws.cell(row=r,column=2).fill=fill(bg); ws.cell(row=r,column=2).alignment=Alignment(horizontal="center")
    for ai, target_age in enumerate(col_ages):
        years = max(0, min(end, target_age) - start + 1)
        fas_sum = round(sum(årsvals.values()) * years)
        life_totals[ai] += fas_sum
        c = ws.cell(row=r, column=3+ai, value=fas_sum if fas_sum > 0 else "-")
        c.font = bf(10, col=("1A1814" if fas_sum > 0 else "CCCCCC"))
        c.fill = fill(bg)
        if isinstance(fas_sum, int): c.number_format = '#,##0'
        c.alignment = Alignment(horizontal="right", vertical="center")
    note = ""
    if fi == 0: note = "BVC, förskola, barnomsorgsavgift subventionerad"
    elif fi == 1: note = "Obligatorisk grundskola, kostnad ca 95 tkr/elev/år (SKR)"
    elif fi == 3: note = "CSN studiemedel + högskoleanslag"
    elif fi == 4: note = "Barnbidrag, föräldrapenning, sjukförsäkring"
    elif fi == 7: note = "Hög sjukvård, äldreomsorg, garantipension"
    c = ws.cell(row=r, column=8, value=note)
    c.font = bf(9, col="666666"); c.fill=fill(bg); c.alignment=Alignment(vertical="center",wrap_text=True)
    ws.row_dimensions[r].height = 26; r += 1

# Totals
ws.cell(row=r, column=1, value="TOTALT LIVSTIDSFÖRMÅNER").font = bf(10, bold=True, col="FFFFFF")
ws.cell(row=r, column=1).fill = fill("1B4D3E")
ws.cell(row=r, column=2, value="0 → ålder").font = bf(9, col="D4EBE0")
ws.cell(row=r, column=2).fill = fill("1B4D3E")
for ai, tot in enumerate(life_totals):
    c = ws.cell(row=r, column=3+ai, value=int(life_totals[ai]))
    c.font = bf(11, bold=True, col="FFFFFF")
    c.fill = fill("1B4D3E")
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[r].height = 28; r += 2

# ─── SEKTION E: Skatt vs förmåner ─────────────────────────────────────────────
ws.merge_cells(f"A{r}:R{r}")
ws[f"A{r}"] = "E. SKATT VS LIVSTIDSFÖRMÅNER — jämförelsetabell (genomsnittsinkomst 500 tkr/år, 2024 priser)"
ws[f"A{r}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ws[f"A{r}"].fill = fill("2E6DA4")
ws[f"A{r}"].alignment = Alignment(indent=1, vertical="center")
ws.row_dimensions[r].height = 24; r += 1

sv_hdrs = ["Åldersprofil", "Livstidsförmåner (tkr)", "Skatt 20–ålder (tkr, 500tkr/år)", "Netto (Förmåner−Skatt)", "Förmåner/Skatt-kvot", "Tolkning"]
sv_widths = [16, 24, 30, 24, 22, 45]
for ci, (h, w) in enumerate(zip(sv_hdrs, sv_widths), 1):
    c = ws.cell(row=r, column=ci, value=h)
    c.font = hf(10)
    c.fill = fill("1A3557")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[gcl(ci)].width = w
ws.row_dimensions[r].height = 36; r += 1

for ai, (ap, lt) in enumerate(zip(AGE_PROFILES, life_totals)):
    skatt = berakna_skatt_livet(ap["age"], 500)
    netto = int(lt) - skatt
    kvot = round(lt / skatt, 2) if skatt > 0 else 0
    if kvot > 1.3:    tolkning = "Tar mer i förmåner än betalar i skatt (hittills)"
    elif kvot > 0.9:  tolkning = "Ungefärlig balans förmåner och skatt"
    else:             tolkning = "Har betalat mer i skatt än fått i förmåner"
    
    netto_col = "1B4D3E" if netto > 0 else "B84A1E"
    bg = "F4F1EB" if (r % 2 == 0) else "FFFFFF"
    for ci, val in enumerate([ap["label"], int(lt), skatt, netto, kvot, tolkning], 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = bf(10, bold=(ci in [1,4,5]), col=(netto_col if ci == 4 else "1A1814"))
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal=("right" if ci in [2,3,4,5] else "left"), vertical="center", wrap_text=(ci==6))
        if ci in [2,3,4]: c.number_format = '#,##0'
        if ci == 5: c.number_format = '0.00"x"'
    ws.row_dimensions[r].height = 28; r += 1
r += 1

# ─── SEKTION F: Metodbegränsningar ────────────────────────────────────────────
ws.merge_cells(f"A{r}:R{r}")
ws[f"A{r}"] = "F. METODBEGRÄNSNINGAR & TRANSPARENSNOTER"
ws[f"A{r}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ws[f"A{r}"].fill = fill("2E6DA4")
ws[f"A{r}"].alignment = Alignment(indent=1, vertical="center")
ws.row_dimensions[r].height = 24; r += 1

begr = [
    ("1. Schabloner ≠ individutfall", "Alla belopp är genomsnitt för åldersgruppen. En frisk 35-åring som aldrig studerat har fått betydligt lägre förmåner. En person med kronisk sjukdom betydligt högre."),
    ("2. Livstidsmodellen är statisk", "Vi håller 2024-prisnivåer konstanta bakåt i tid. Realt var sjukvård billigare 1990, skolan dyrare per elev. En korrekt livstidsmodell kräver dynamisk historisk fördelning."),
    ("3. Inkomstpension ingår ej", "Inkomstpensionen (ATP) är avgiftsfinansierad — en egen post. Garantipensionen (UO11) ingår. Totalpensionen underskattas för de flesta."),
    ("4. Kommunal vs statlig nivå", "Kommunal skola och sjukvård via regioner finansieras av kommunalskatt, ej statsbudget. Vi inkluderar SKR-data för fullständig bild men gränsdragning är metodologisk."),
    ("5. Vikter baseras på aggregerad statistik", "Åldersviktarna bygger på gruppaggregat från Socialstyrelsen, FK, CSN m.fl. Individuella variationer fångas ej. Vikterna bör uppdateras vartannat år."),
    ("6. Prisbasårseffekter", "Alla siffror i 2024 priser via SCB KPI. Relativa prisförändringar inom sektorer (sjukvård, utbildning) jämfört med KPI fångas ej — dessa sektorer har haft högre prisutveckling."),
]

for nr, (titel, text) in enumerate(begr):
    bg = "FFF9E6" if (nr % 2 == 0) else "F4F1EB"
    ws.cell(row=r, column=1, value=titel).font = bf(10, bold=True)
    ws.cell(row=r, column=1).fill = fill(bg)
    ws.merge_cells(f"B{r}:R{r}")
    ws.cell(row=r, column=2, value=text).font = bf(9, col="444444")
    ws.cell(row=r, column=2).fill = fill(bg)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 36; r += 1

# Column widths for col A
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 11
ws.column_dimensions["D"].width = 11
ws.column_dimensions["E"].width = 11
ws.column_dimensions["F"].width = 11
ws.column_dimensions["G"].width = 11
ws.column_dimensions["H"].width = 50

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(_EXCEL)

# ── Export JSON for website ───────────────────────────────────────────────────
model_export = {
    "basår": BASÅR,
    "befolkning": BEFOLKNING,
    "per_capita_tkr": PC,
    "vikter": VIKTER,
    "poster": [{"namn": n, "key": k, "källa": s} for n,k,s in POSTER_DEF],
    "age_profiles": [
        {
            "label": ap["label"],
            "age": ap["age"],
            "årsförmåner": berakna_arsformaaner(ap["idx"]),
            "livstid_tkr": int(life_totals[ap["idx"]]),
            "skatt_500tkr": berakna_skatt_livet(ap["age"], 500),
        }
        for ap in AGE_PROFILES
    ],
    "livsfaser": [
        {"label": lbl, "start": s, "end": e, "årsvals": v}
        for (s,e,v), lbl in zip(LIVSFASER, LIVSFAS_LABELS)
    ],
}

with open(os.path.join(_DIR, 'nta_profil_v2.json'), 'w', encoding='utf-8') as f:
    json.dump(model_export, f, ensure_ascii=False, indent=2)

print("✓ Excel workbook saved")
print("✓ JSON model exported")
print("\n=== SAMMANFATTNING ===")
for ap in AGE_PROFILES:
    lt = int(life_totals[ap["idx"]])
    sk = berakna_skatt_livet(ap["age"], 500)
    kvot = round(lt/sk, 2) if sk > 0 else "n/a"
    print(f"  {ap['label']:8s}  Livstidsförmåner: {lt:>6,} tkr  |  Skatt 20→{ap['age']}: {sk:>5,} tkr  |  Kvot: {kvot}x")

# ── Verifieringstest (kör med: python nta_berakning.py --verify) ─────────────
if '--verify' in sys.argv:
    print("\n=== VERIFIERINGSTEST ===")
    errors = []

    # Test 1: livsfaser_v2.json laddas korrekt
    assert len(LIVSFASER) == 10, f"Fel: förväntade 10 faser, fick {len(LIVSFASER)}"
    print("  ✓ livsfaser_v2.json laddad: 10 faser")

    # Test 2: Alla faser täcker 0–85 utan gap
    prev_end = -1
    for i, (start, end, _) in enumerate(LIVSFASER):
        assert start == prev_end + 1, f"Gap i fas {i}: start={start}, prev_end={prev_end}"
        prev_end = end
    print(f"  ✓ Faserna täcker 0–{prev_end} utan luckor")

    # Test 3: Totala livstidsförmåner rimliga (300–20 000 tkr per profil)
    for ap in AGE_PROFILES:
        lt = int(life_totals[ap["idx"]])
        assert 100 < lt < 25000, f"Orimligt livstidsvärde för {ap['label']}: {lt} tkr"
    print("  ✓ Livstidsförmåner inom rimliga gränser för alla profiler")

    # Test 4: Skatteberäkning rimlig (exkl pension, 500 tkr/år)
    skatt_35 = berakna_skatt_livet(35, 500)
    assert 5000 < skatt_35 < 30000, f"Orimlig skatt vid 35 år: {skatt_35} tkr"
    print(f"  ✓ Skatteberäkning rimlig: 35 år, 500 tkr/år → {skatt_35:,} tkr ackumulerat")

    # Test 5: Kalibrering mot COFOG (max 20% avvikelse)
    BEFOLKNING_PER_FAS = {
        0:700000, 1:1000000, 2:350000, 3:750000, 4:1200000,
        5:1400000, 6:1300000, 7:1200000, 8:1100000, 9:800000
    }
    livsfaser_path = os.path.join(_DIR, 'livsfaser_v2.json')
    with open(livsfaser_path, encoding='utf-8') as fh:
        lf_data = json.load(fh)
    total_mdkr = sum(
        BEFOLKNING_PER_FAS[i] * f['total_tkr_per_år'] / 1_000_000_000 * 1_000
        for i, f in enumerate(lf_data['faser'])
    )
    cofog_target = 3011
    diff_pct = abs(total_mdkr - cofog_target) / cofog_target * 100
    assert diff_pct < 20, f"COFOG-differens för stor: {diff_pct:.1f}%"
    print(f"  ✓ COFOG-kalibrering: {total_mdkr:.0f} Mdkr vs mål {cofog_target} Mdkr ({diff_pct:.1f}% differens)")

    if errors:
        print(f"\n  ✗ {len(errors)} FEL:")
        for e in errors: print(f"    - {e}")
        sys.exit(1)
    else:
        print("\n  ✓ Alla verifieringstester godkända")
